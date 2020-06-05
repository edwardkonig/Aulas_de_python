import xlsxwriter
import openpyxl
# workbook = xlsxwriter.Workbook('dictExcel.xlsx')
# worksheet = workbook.add_worksheet()
dict_perguntas = {
  "nome":  "Qual é o seu nome?", 
   "idade": "Qual é a sua idade?",
   "sexo": "Qual é o seu sexo?",
   "cidade": "Onde você mora?"
}

dict_respostas = dict.fromkeys(dict_perguntas.keys(), "") # e se fosse um valor com string cheia?

for k,v in dict_perguntas.items():
    r = input(v)
    dict_respostas[k]= r
# def mergeDict(dict_perguntas, dict_respostas):
#    dict_combined = {**dict_perguntas, **dict_respostas}
#    for key, value in dict_combined.items():
#        if key in dict_perguntas and key in dict_respostas:
#                dict_combined[key] = [value , dict_perguntas[key]]
 
#    return dict_combined
 
# dict_combined = mergeDict(dict_perguntas, dict_respostas)
wb = openpyxl.load_workbook(filename='dictExcel.xlsx')
ws = wb.active
row = 1
col = 1

for key,value in dict_respostas.items():
    ws.cell(row, col, key)
    ws.cell(row, col+1, value)
    row+=1
for row in dict_respostas:
    ws.append(row)

    
wb.save('dictExcel.xlsx')
# workbook.close()

'''
https://thispointer.com/how-to-merge-two-or-more-dictionaries-in-python/
https://stackoverflow.com/questions/33575376/using-user-input-to-create-dictionaries-in-python
https://stackoverflow.com/questions/45201288/how-to-create-header-in-excel-from-a-python-dictionary-keys
'''
'''
Make Excel read + write
Study list comprehension
Rotate list and keep only one key
'''

